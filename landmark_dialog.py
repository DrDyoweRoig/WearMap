"""
2D Landmarking & Procrustes Cluster Analysis module for WearMap.

Workflow:
  1. Load images and place N landmarks on each (left-click to add, right-click to undo).
  2. Click "Run Analysis": runs Generalized Procrustes Analysis (GPA) on all complete
     landmark sets, then k-means clustering on the aligned shapes.
  3. Results show cluster assignment, centroid size, and Procrustes distance from mean.
  4. Export results to Excel or CSV (summary + aligned coordinates).
"""

import math
import os

import numpy as np
from PyQt6.QtCore import Qt, QRectF, pyqtSignal
from PyQt6.QtGui import (
    QBrush, QColor, QFont, QImageReader, QPainter, QPen, QPixmap,
)
from PyQt6.QtWidgets import (
    QCheckBox, QDialog, QFileDialog, QFrame, QGroupBox, QHBoxLayout,
    QHeaderView, QLabel, QListWidget, QListWidgetItem, QMessageBox,
    QPushButton, QSpinBox, QSplitter, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QWidget, QGraphicsView, QGraphicsScene,
)


# ── Colours ──────────────────────────────────────────────────────────────────

_LM_COLORS = [
    "#EF4444", "#3B82F6", "#22C55E", "#F59E0B", "#8B5CF6",
    "#EC4899", "#14B8A6", "#F97316", "#84CC16", "#06B6D4",
]
_CL_COLORS = [
    "#EF4444", "#3B82F6", "#22C55E", "#F59E0B", "#8B5CF6",
    "#EC4899", "#14B8A6", "#F97316",
]


# ── Maths ─────────────────────────────────────────────────────────────────────

def _gpa(shapes: list) -> tuple:
    """
    Generalized Procrustes Analysis (numpy-only, no scipy required).
    shapes: list of np.ndarray(n, 2)
    Returns (aligned_shapes, mean_shape) — each shape is centred, unit-size, rotated.
    """
    shapes = [s.astype(float).copy() for s in shapes]

    # 1. Translate to centroid
    for i in range(len(shapes)):
        shapes[i] -= shapes[i].mean(axis=0)

    # 2. Scale to unit centroid size
    for i in range(len(shapes)):
        cs = math.sqrt(float((shapes[i] ** 2).sum()))
        if cs > 0:
            shapes[i] /= cs

    # 3. Iterative Procrustes rotation alignment
    for _ in range(300):
        mean_shape = np.mean(shapes, axis=0)
        ms_cs = math.sqrt(float((mean_shape ** 2).sum()))
        if ms_cs > 0:
            mean_shape /= ms_cs

        changed = False
        for i in range(len(shapes)):
            M = shapes[i].T @ mean_shape
            U, _, Vt = np.linalg.svd(M)
            R = Vt.T @ U.T
            new_s = shapes[i] @ R.T
            if not np.allclose(new_s, shapes[i], atol=1e-10):
                changed = True
            shapes[i] = new_s

        if not changed:
            break

    mean_shape = np.mean(shapes, axis=0)
    return shapes, mean_shape


def _kmeans(X: np.ndarray, k: int, n_init: int = 10, max_iter: int = 300) -> np.ndarray:
    """K-means++ clustering. Returns integer labels array of length len(X)."""
    n = len(X)
    if k >= n:
        return np.arange(n, dtype=int)

    rng = np.random.default_rng(42)
    best_labels = np.zeros(n, dtype=int)
    best_inertia = float("inf")

    for _ in range(n_init):
        # k-means++ initialisation
        idx = [int(rng.integers(n))]
        for _ in range(k - 1):
            dists = np.array([
                min(float(np.sum((X[i] - X[ci]) ** 2)) for ci in idx)
                for i in range(n)
            ])
            total = dists.sum()
            probs = dists / total if total > 0 else np.ones(n) / n
            idx.append(int(rng.choice(n, p=probs)))
        centers = X[idx].copy().astype(float)

        labels = np.zeros(n, dtype=int)
        for _ in range(max_iter):
            dists = np.array([[float(np.sum((x - c) ** 2)) for c in centers] for x in X])
            new_labels = np.argmin(dists, axis=1)
            if np.all(new_labels == labels):
                break
            labels = new_labels
            for j in range(k):
                mask = labels == j
                if mask.any():
                    centers[j] = X[mask].mean(axis=0)

        inertia = sum(float(np.sum((X[i] - centers[labels[i]]) ** 2)) for i in range(n))
        if inertia < best_inertia:
            best_inertia = inertia
            best_labels = labels.copy()

    return best_labels


def _auto_k(X: np.ndarray, max_k: int = 8) -> int:
    """Elbow method: return the k at which the WCSS curve bends most sharply."""
    n = len(X)
    max_k = min(max_k, n - 1)
    if max_k < 2:
        return 2

    wcss = []
    for k in range(1, max_k + 1):
        if k == 1:
            mean = X.mean(axis=0)
            wcss.append(float(np.sum((X - mean) ** 2)))
        else:
            labels = _kmeans(X, k, n_init=5)
            inertia = 0.0
            for j in range(k):
                mask = labels == j
                if mask.any():
                    c = X[mask].mean(axis=0)
                    inertia += float(np.sum((X[mask] - c) ** 2))
            wcss.append(inertia)

    if len(wcss) < 3:
        return 2

    best_k, best_dd = 2, 0.0
    for i in range(1, len(wcss) - 1):
        dd = wcss[i - 1] - 2 * wcss[i] + wcss[i + 1]
        if dd > best_dd:
            best_dd = dd
            best_k = i + 1  # k is 1-indexed but loop is 0-indexed starting from k=1

    return max(2, best_k)


# ── Landmark canvas ───────────────────────────────────────────────────────────

class LandmarkCanvas(QGraphicsView):
    """Mini image view where the user clicks to place numbered landmarks."""

    landmark_changed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setBackgroundBrush(QBrush(QColor("#1e2030")))
        self.setStyleSheet("border: none;")

        self._pixmap_item = None
        self._lm_items: list = []
        self._landmarks: list = []   # [(x, y), ...]
        self._max_n: int = 8

    # ── Public API ────────────────────────────────────────────────────────

    def load_image(self, path: str) -> bool:
        reader = QImageReader(path)
        reader.setAutoTransform(True)
        img = reader.read()
        if img.isNull():
            return False
        px = QPixmap.fromImage(img)
        self._scene.clear()
        self._lm_items.clear()
        self._pixmap_item = self._scene.addPixmap(px)
        self.setSceneRect(QRectF(px.rect()))
        self.fitInView(self._pixmap_item, Qt.AspectRatioMode.KeepAspectRatio)
        return True

    def set_landmarks(self, lms: list, max_n: int):
        self._max_n = max_n
        self._landmarks = list(lms)
        self._redraw()

    def set_max_n(self, n: int):
        self._max_n = n
        if len(self._landmarks) > n:
            self._landmarks = self._landmarks[:n]
            self._redraw()
            self.landmark_changed.emit()

    def get_landmarks(self) -> list:
        return list(self._landmarks)

    def clear_landmarks(self):
        self._landmarks.clear()
        self._redraw()
        self.landmark_changed.emit()

    # ── Mouse events ──────────────────────────────────────────────────────

    def wheelEvent(self, event):
        factor = 1.15 if event.angleDelta().y() > 0 else 1 / 1.15
        self.scale(factor, factor)

    def mousePressEvent(self, event):
        if self._pixmap_item is None:
            super().mousePressEvent(event)
            return
        if event.button() == Qt.MouseButton.LeftButton:
            if len(self._landmarks) < self._max_n:
                sp = self.mapToScene(event.pos())
                self._landmarks.append((sp.x(), sp.y()))
                self._redraw()
                self.landmark_changed.emit()
            return
        if event.button() == Qt.MouseButton.RightButton:
            if self._landmarks:
                self._landmarks.pop()
                self._redraw()
                self.landmark_changed.emit()
            return
        super().mousePressEvent(event)

    # ── Drawing ───────────────────────────────────────────────────────────

    def _redraw(self):
        for item in self._lm_items:
            self._scene.removeItem(item)
        self._lm_items.clear()

        font = QFont("Segoe UI", 8)
        font.setBold(True)

        for i, (x, y) in enumerate(self._landmarks):
            color = QColor(_LM_COLORS[i % len(_LM_COLORS)])
            r = 5.0
            pen = QPen(color, 1.5)
            pen.setCosmetic(True)
            fill = QColor(color)
            fill.setAlpha(200)
            ellipse = self._scene.addEllipse(x - r, y - r, 2 * r, 2 * r, pen, QBrush(fill))
            self._lm_items.append(ellipse)

            text = self._scene.addSimpleText(str(i + 1))
            text.setFont(font)
            text.setPos(x + 6, y - 8)
            text.setBrush(QBrush(color))
            self._lm_items.append(text)


# ── Main dialog ───────────────────────────────────────────────────────────────

class LandmarkingDialog(QDialog):
    """2D Landmarking & Procrustes Cluster Analysis."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("2D Landmarking & Procrustes Cluster Analysis")
        self.resize(1100, 740)
        self.setModal(False)

        # sessions: [{"path": str, "name": str, "landmarks": [(x,y),...]}]
        self._sessions: list = []
        self._current_idx: int = -1
        self._analysis_results: dict | None = None

        self._build_ui()

    # ── UI construction ───────────────────────────────────────────────────

    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setSpacing(6)
        main.setContentsMargins(8, 8, 8, 8)

        # ── Top: list panel + landmark canvas ──────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ── Left: session list ────────────────────────────────────────────
        left = QWidget()
        left.setMaximumWidth(260)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(4, 4, 4, 4)
        ll.setSpacing(6)

        hdr = QLabel("Images")
        hdr.setStyleSheet("font-weight:700; color:#14532D; font-size:13px;")
        ll.addWidget(hdr)

        self._list = QListWidget()
        self._list.currentRowChanged.connect(self._on_row_changed)
        ll.addWidget(self._list)

        brow = QHBoxLayout()
        add_btn = QPushButton("+ Add")
        add_btn.setToolTip("Add image(s) to landmark")
        add_btn.clicked.connect(self._add_images)
        rem_btn = QPushButton("Remove")
        rem_btn.clicked.connect(self._remove_current)
        brow.addWidget(add_btn)
        brow.addWidget(rem_btn)
        ll.addLayout(brow)

        clr_btn = QPushButton("Clear All")
        clr_btn.clicked.connect(self._clear_all)
        ll.addWidget(clr_btn)

        ll.addWidget(self._sep())

        nrow = QHBoxLayout()
        nrow.addWidget(QLabel("N landmarks:"))
        self._n_spin = QSpinBox()
        self._n_spin.setRange(3, 500)
        self._n_spin.setValue(8)
        self._n_spin.setToolTip("Expected number of landmarks per image")
        self._n_spin.valueChanged.connect(self._on_n_changed)
        nrow.addWidget(self._n_spin)
        ll.addLayout(nrow)

        self._status_lbl = QLabel("No images loaded.")
        self._status_lbl.setStyleSheet("color:#4B7A5A; font-size:11px;")
        self._status_lbl.setWordWrap(True)
        ll.addWidget(self._status_lbl)
        ll.addStretch()

        splitter.addWidget(left)

        # ── Right: canvas + info bar ──────────────────────────────────────
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(4, 4, 4, 4)
        rl.setSpacing(4)

        self._canvas = LandmarkCanvas()
        self._canvas.landmark_changed.connect(self._on_landmark_changed)
        rl.addWidget(self._canvas)

        info_row = QHBoxLayout()
        self._lm_lbl = QLabel("Select an image from the list to start.")
        self._lm_lbl.setStyleSheet("color:#166534; font-size:12px;")
        info_row.addWidget(self._lm_lbl)
        info_row.addStretch()
        reset_btn = QPushButton("Reset Landmarks")
        reset_btn.setToolTip("Remove all landmarks for this image")
        reset_btn.clicked.connect(self._reset_landmarks)
        info_row.addWidget(reset_btn)
        rl.addLayout(info_row)

        splitter.addWidget(right)
        splitter.setSizes([240, 840])
        main.addWidget(splitter, stretch=3)

        # ── Bottom: analysis controls ─────────────────────────────────────
        main.addWidget(self._sep())

        ctrl = QHBoxLayout()
        ctrl.addWidget(QLabel("Clusters (k):"))
        self._k_spin = QSpinBox()
        self._k_spin.setRange(2, 20)
        self._k_spin.setValue(3)
        self._k_spin.setEnabled(False)
        ctrl.addWidget(self._k_spin)

        self._auto_k_chk = QCheckBox("Auto-detect k (elbow method)")
        self._auto_k_chk.setChecked(True)
        self._auto_k_chk.toggled.connect(lambda checked: self._k_spin.setEnabled(not checked))
        ctrl.addWidget(self._auto_k_chk)
        ctrl.addStretch()

        run_btn = QPushButton("▶  Run Analysis")
        run_btn.setStyleSheet(
            "font-weight:bold; padding:5px 18px;"
            "background:#22C55E; color:#14532D; border-radius:4px;"
        )
        run_btn.setToolTip(
            "Run Generalized Procrustes Analysis + k-means clustering\n"
            "on all images that have the correct number of landmarks."
        )
        run_btn.clicked.connect(self._run_analysis)
        ctrl.addWidget(run_btn)
        main.addLayout(ctrl)

        # ── Results table ─────────────────────────────────────────────────
        grp = QGroupBox("Results")
        grp.setStyleSheet(
            "QGroupBox{font-weight:700;color:#14532D;"
            "border:1px solid #86EFAC;border-radius:6px;"
            "margin-top:6px;padding:6px;}"
            "QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 4px;}"
        )
        grp_l = QVBoxLayout(grp)

        self._results_table = QTableWidget(0, 5)
        self._results_table.setHorizontalHeaderLabels(
            ["Image", "Cluster", "Centroid Size", "Procrustes Dist.", "N Landmarks"]
        )
        self._results_table.setAlternatingRowColors(True)
        self._results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._results_table.verticalHeader().setVisible(False)
        hh = self._results_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for col in range(1, 5):
            hh.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
        grp_l.addWidget(self._results_table)

        exp_row = QHBoxLayout()
        exp_row.addStretch()
        exp_btn = QPushButton("Export Results…")
        exp_btn.clicked.connect(self._export_results)
        exp_row.addWidget(exp_btn)
        grp_l.addLayout(exp_row)

        main.addWidget(grp, stretch=1)

    @staticmethod
    def _sep() -> QFrame:
        f = QFrame()
        f.setFrameShape(QFrame.Shape.HLine)
        f.setStyleSheet("background:#BBF7D0; max-height:1px; border:none;")
        return f

    # ── Session management ────────────────────────────────────────────────

    def _add_images(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Add Images", "",
            "Images (*.png *.jpg *.jpeg *.tif *.tiff *.bmp *.webp *.plu *.plux);;All (*)",
        )
        added = 0
        for p in paths:
            if any(s["path"] == p for s in self._sessions):
                continue
            self._sessions.append({"path": p, "name": os.path.basename(p), "landmarks": []})
            item = QListWidgetItem("")
            self._list.addItem(item)
            self._refresh_item(len(self._sessions) - 1)
            added += 1
        if added and self._current_idx < 0:
            self._list.setCurrentRow(0)
        self._update_status()

    def _remove_current(self):
        row = self._list.currentRow()
        if row < 0:
            return
        self._sessions.pop(row)
        self._list.takeItem(row)
        if self._sessions:
            self._list.setCurrentRow(min(row, len(self._sessions) - 1))
        else:
            self._current_idx = -1
            self._update_lm_label()
        self._update_status()

    def _clear_all(self):
        self._sessions.clear()
        self._list.clear()
        self._current_idx = -1
        self._results_table.setRowCount(0)
        self._analysis_results = None
        self._update_lm_label()
        self._update_status()

    def _refresh_item(self, idx: int):
        if not (0 <= idx < len(self._sessions)):
            return
        s = self._sessions[idx]
        n = self._n_spin.value()
        placed = len(s["landmarks"])
        marker = "✓" if placed == n else f"{placed}/{n}"
        self._list.item(idx).setText(f"{marker}  {s['name']}")

    def _on_row_changed(self, row: int):
        # Save landmarks of the previously active session
        if 0 <= self._current_idx < len(self._sessions):
            self._sessions[self._current_idx]["landmarks"] = self._canvas.get_landmarks()
            self._refresh_item(self._current_idx)

        self._current_idx = row
        if not (0 <= row < len(self._sessions)):
            self._update_lm_label()
            return

        s = self._sessions[row]
        if not self._canvas.load_image(s["path"]):
            self._lm_lbl.setText(f"Could not load: {s['name']}")
            return
        self._canvas.set_landmarks(s["landmarks"], self._n_spin.value())
        self._update_lm_label()

    def _on_landmark_changed(self):
        if 0 <= self._current_idx < len(self._sessions):
            self._sessions[self._current_idx]["landmarks"] = self._canvas.get_landmarks()
            self._refresh_item(self._current_idx)
        self._update_lm_label()
        self._update_status()

    def _reset_landmarks(self):
        if 0 <= self._current_idx < len(self._sessions):
            self._sessions[self._current_idx]["landmarks"] = []
        self._canvas.clear_landmarks()
        self._refresh_item(self._current_idx)
        self._update_lm_label()
        self._update_status()

    def _on_n_changed(self, n: int):
        self._canvas.set_max_n(n)
        for i in range(len(self._sessions)):
            if len(self._sessions[i]["landmarks"]) > n:
                self._sessions[i]["landmarks"] = self._sessions[i]["landmarks"][:n]
            self._refresh_item(i)
        self._update_lm_label()
        self._update_status()

    def _update_lm_label(self):
        if self._current_idx < 0 or not self._sessions:
            self._lm_lbl.setText("Select an image from the list to start.")
            return
        n = self._n_spin.value()
        placed = len(self._canvas.get_landmarks())
        remaining = n - placed
        if remaining > 0:
            self._lm_lbl.setText(
                f"Landmarks: {placed}/{n}  —  Left-click to place · Right-click to undo"
            )
        else:
            self._lm_lbl.setText(f"Landmarks: {placed}/{n}  ✓ Complete")

    def _update_status(self):
        n = self._n_spin.value()
        total = len(self._sessions)
        complete = sum(1 for s in self._sessions if len(s["landmarks"]) == n)
        if total == 0:
            self._status_lbl.setText("No images loaded.")
        else:
            self._status_lbl.setText(
                f"{total} image(s) · {complete}/{total} complete\n({n} landmarks each)"
            )

    # ── Analysis ──────────────────────────────────────────────────────────

    def _run_analysis(self):
        # Commit current canvas landmarks
        if 0 <= self._current_idx < len(self._sessions):
            self._sessions[self._current_idx]["landmarks"] = self._canvas.get_landmarks()
            self._refresh_item(self._current_idx)

        n = self._n_spin.value()
        complete = [s for s in self._sessions if len(s["landmarks"]) == n]

        if len(complete) < 3:
            QMessageBox.warning(
                self, "Not enough data",
                f"Need at least 3 images with {n} landmarks each.\n"
                f"Currently {len(complete)} complete.",
            )
            return

        shapes = [np.array(s["landmarks"], dtype=float) for s in complete]

        # Centroid sizes (before GPA normalisation)
        cs_list = []
        for s in shapes:
            c = s - s.mean(axis=0)
            cs_list.append(float(np.sqrt((c ** 2).sum())))

        # Generalized Procrustes Analysis
        aligned, mean_shape = _gpa(shapes)

        # Procrustes distances from grand mean
        proc_dists = [float(np.sqrt(((a - mean_shape) ** 2).sum())) for a in aligned]

        # Cluster on flattened aligned shapes
        X = np.array([a.flatten() for a in aligned])
        if self._auto_k_chk.isChecked():
            k = _auto_k(X)
            self._k_spin.setValue(k)
        else:
            k = self._k_spin.value()
        k = max(2, min(k, len(complete)))

        labels = _kmeans(X, k)

        self._analysis_results = {
            "sessions": complete,
            "aligned": aligned,
            "labels": labels,
            "proc_dists": proc_dists,
            "cs_list": cs_list,
            "mean_shape": mean_shape,
            "k": int(k),
            "n": n,
        }
        self._fill_results_table()

    def _fill_results_table(self):
        res = self._analysis_results
        if not res:
            return
        self._results_table.setRowCount(0)
        for i, s in enumerate(res["sessions"]):
            row = self._results_table.rowCount()
            self._results_table.insertRow(row)
            cl = int(res["labels"][i]) + 1
            color = QColor(_CL_COLORS[(cl - 1) % len(_CL_COLORS)])
            for col, txt in enumerate([
                s["name"],
                f"C{cl}",
                f"{res['cs_list'][i]:.4f}",
                f"{res['proc_dists'][i]:.6f}",
                str(res["n"]),
            ]):
                item = QTableWidgetItem(txt)
                item.setForeground(color)
                self._results_table.setItem(row, col, item)

    # ── Export ────────────────────────────────────────────────────────────

    def _export_results(self):
        if not self._analysis_results:
            QMessageBox.information(self, "Export", "Run the analysis first.")
            return

        path, sel = QFileDialog.getSaveFileName(
            self, "Export Landmark Results", "landmark_results",
            "Excel (*.xlsx);;CSV (*.csv)",
        )
        if not path:
            return

        res = self._analysis_results
        n = res["n"]

        # Summary sheet
        summ_hdrs = ["Image", "Cluster", "Centroid Size", "Procrustes Dist.", "N Landmarks"]
        summ_rows = [
            [
                s["name"],
                int(res["labels"][i]) + 1,
                f"{res['cs_list'][i]:.8f}",
                f"{res['proc_dists'][i]:.8f}",
                n,
            ]
            for i, s in enumerate(res["sessions"])
        ]

        # Aligned coordinates sheet
        coord_cols = []
        for j in range(n):
            coord_cols += [f"LM{j+1}_X", f"LM{j+1}_Y"]
        coord_hdrs = ["Image", "Cluster"] + coord_cols + ["Centroid Size", "Procrustes Dist."]
        coord_rows = []
        for i, s in enumerate(res["sessions"]):
            cl = int(res["labels"][i]) + 1
            flat = []
            for (x, y) in res["aligned"][i]:
                flat += [f"{x:.8f}", f"{y:.8f}"]
            coord_rows.append(
                [s["name"], cl] + flat +
                [f"{res['cs_list'][i]:.8f}", f"{res['proc_dists'][i]:.8f}"]
            )

        try:
            if "xlsx" in path or "xlsx" in sel:
                if not path.endswith(".xlsx"):
                    path += ".xlsx"
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
                wb = openpyxl.Workbook()
                hf = Font(bold=True, color="14532D")
                hfill = PatternFill("solid", fgColor="86EFAC")

                def _fill(ws, hdrs, rows):
                    for ci, h in enumerate(hdrs, 1):
                        cell = ws.cell(row=1, column=ci, value=h)
                        cell.font = hf
                        cell.fill = hfill
                        cell.alignment = Alignment(horizontal="center")
                    for ri, r in enumerate(rows, 2):
                        for ci, v in enumerate(r, 1):
                            ws.cell(row=ri, column=ci, value=v)
                    for col_cells in ws.columns:
                        w = max(len(str(c.value or "")) for c in col_cells) + 4
                        ws.column_dimensions[col_cells[0].column_letter].width = min(w, 40)

                ws1 = wb.active
                ws1.title = "Summary"
                _fill(ws1, summ_hdrs, summ_rows)
                ws2 = wb.create_sheet("Aligned Coords")
                _fill(ws2, coord_hdrs, coord_rows)
                wb.save(path)
            else:
                if not path.endswith(".csv"):
                    path += ".csv"
                import csv
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow(coord_hdrs)
                    for r in coord_rows:
                        w.writerow(r)

            QMessageBox.information(self, "Export", f"Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
