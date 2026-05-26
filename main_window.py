import os
import sys

from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import (
    QAction, QColor, QDragEnterEvent, QDropEvent,
    QKeySequence, QShortcut,
)
from PyQt6.QtGui import QPixmap
from PyQt6.QtWidgets import (
    QCheckBox, QComboBox, QDialog, QFileDialog, QFrame, QGridLayout,
    QGroupBox, QHeaderView, QHBoxLayout, QLabel, QMainWindow, QMessageBox,
    QMenu, QPushButton, QSlider, QSplitter, QTableWidget, QTableWidgetItem,
    QTabWidget, QToolButton, QVBoxLayout, QWidget,
)

from canvas import ImageCanvas, Tool, PUECH_COLORS
from dialogs import AboutDialog, CalibrationDialog, ScaleBarDialog, StriationTypeNamesDialog
from landmark_dialog import LandmarkingDialog
from measurements import MeasurementType
from style import MAIN_STYLE

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".webp", ".plu", ".plux"}


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Wearmap \u2014 Image Analysis")
        self.setMinimumSize(1180, 740)
        self.setAcceptDrops(True)
        self.setStyleSheet(MAIN_STYLE)

        self.canvas = ImageCanvas()
        self.canvas.measurement_added.connect(self._on_measurement_added)
        self.canvas.measurement_clicked.connect(self._on_canvas_measurement_clicked)
        self.canvas.calibration_line_drawn.connect(self._on_calibration_line)
        self.canvas.position_changed.connect(self._update_position)

        # Multi-image navigation state
        self._image_files: list[str] = []
        self._image_idx: int = -1
        self._image_states: dict[str, dict] = {}   # path → canvas state dict
        self._current_path: str = ""

        self._special_filter: str | None = None
        self._stype_names: dict = {1: "Fine", 2: "Medium", 3: "Coarse"}
        self._landmarking_dlg: LandmarkingDialog | None = None
        self._stype_target_mids: set[int] = set()  # measurement IDs targeted for stype

        self._build_ui()
        self._build_menu()
        self._build_toolbar()
        self._build_statusbar()
        self._build_shortcuts()

    # ── Layout ────────────────────────────────────────────────────────────

    def _build_ui(self):
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.canvas)
        splitter.addWidget(self._build_panel())
        splitter.setSizes([780, 340])
        splitter.setChildrenCollapsible(False)
        self.setCentralWidget(splitter)

    def _build_panel(self) -> QWidget:
        panel = QWidget()
        panel.setObjectName("sidePanel")
        panel.setMinimumWidth(280)
        tabs = QTabWidget()
        tabs.addTab(self._build_measures_tab(), "Measurements")
        tabs.addTab(self._build_filters_tab(), "Filters")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.addWidget(tabs)
        return panel

    # ── Measurements tab ──────────────────────────────────────────────────

    def _build_measures_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(8, 10, 8, 8)
        layout.setSpacing(8)

        layout.addWidget(self._build_tooth_context())

        # Table: #, Type, Value, Slope, Puech, SType, Label
        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["#", "Type", "Value", "Slope", "Puech", "SType", "Label"]
        )
        hh = self.table.horizontalHeader()
        for col in range(6):
            hh.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.itemSelectionChanged.connect(self._on_table_selection_changed)
        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        self.del_btn = QPushButton("Delete")
        self.del_btn.setObjectName("deleteBtn")
        self.del_btn.setToolTip("Delete selected measurement (Del)")
        self.del_btn.clicked.connect(self._delete_selected)
        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setObjectName("deleteBtn")
        self.clear_btn.clicked.connect(self._clear_all)
        btn_row.addWidget(self.del_btn)
        btn_row.addWidget(self.clear_btn)
        layout.addLayout(btn_row)

        # Striation type panel (hidden by default, toggled from Analysis menu)
        self._stype_panel = QWidget()
        stype_v = QVBoxLayout(self._stype_panel)
        stype_v.setContentsMargins(0, 2, 0, 2)
        stype_v.setSpacing(4)
        stype_row = QHBoxLayout()
        stype_row.addWidget(QLabel("Striation type:"))
        self._stype_combo = QComboBox()
        self._stype_combo.setToolTip(
            "Assign striation type to selected measurement(s).\n"
            "Shortcuts: 1 / 2 / 3 … assign by type number.\n"
            "Use Analysis → Rename Types… to customise names."
        )
        self._rebuild_stype_combo()
        self._stype_combo.currentIndexChanged.connect(self._on_stype_combo_changed)
        stype_row.addWidget(self._stype_combo)
        stype_v.addLayout(stype_row)
        stype_hint = QLabel(
            "ℹ  Workflow: draw a line first, then press 1 / 2 / 3 on the "
            "keyboard to assign its striation type."
        )
        stype_hint.setWordWrap(True)
        stype_hint.setStyleSheet(
            "color:#92400E; background:#FEF3C7; border-radius:4px; "
            "padding:6px; font-size:11px;"
        )
        stype_v.addWidget(stype_hint)
        self._stype_panel.setVisible(False)
        layout.addWidget(self._stype_panel)

        layout.addWidget(self._sep())

        lbl2 = QLabel("Image Info")
        lbl2.setObjectName("panelTitle")
        layout.addWidget(lbl2)
        self.info_label = QLabel("No image loaded.\nDrag & drop or use File \u2192 Open.")
        self.info_label.setWordWrap(True)
        self.info_label.setStyleSheet("color: #4B7A5A; font-size: 12px;")
        layout.addWidget(self.info_label)

        layout.addWidget(self._sep())

        lbl3 = QLabel("Calibration")
        lbl3.setObjectName("panelTitle")
        layout.addWidget(lbl3)
        self.cal_label = QLabel("Not calibrated (px).\nUse \u229e Calibrate to set scale.")
        self.cal_label.setWordWrap(True)
        self.cal_label.setStyleSheet("color: #4B7A5A; font-size: 12px;")
        layout.addWidget(self.cal_label)

        layout.addStretch()

        # Logo bottom-right
        _base = sys._MEIPASS if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
        _logo_path = os.path.join(_base, "logo.png")
        _logo_lbl = QLabel()
        _logo_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        if os.path.exists(_logo_path):
            _px = QPixmap(_logo_path)
            if not _px.isNull():
                _logo_lbl.setPixmap(_px.scaled(
                    82, 82,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                ))
        layout.addWidget(_logo_lbl)
        return w

    # ── Filters tab ───────────────────────────────────────────────────────

    def _build_filters_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(10, 12, 10, 10)
        layout.setSpacing(10)

        def slider_row(label: str, lo: int, hi: int, default: int) -> QSlider:
            row = QHBoxLayout()
            lbl = QLabel(label)
            lbl.setFixedWidth(86)
            sl = QSlider(Qt.Orientation.Horizontal)
            sl.setRange(lo, hi)
            sl.setValue(default)
            val = QLabel(str(default))
            val.setFixedWidth(32)
            val.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            sl.valueChanged.connect(lambda v, lv=val: lv.setText(str(v)))
            sl.valueChanged.connect(lambda _: self._apply_filters())
            row.addWidget(lbl)
            row.addWidget(sl)
            row.addWidget(val)
            layout.addLayout(row)
            return sl

        lbl_adj = QLabel("Adjustments")
        lbl_adj.setObjectName("panelTitle")
        layout.addWidget(lbl_adj)

        self._sl_brightness = slider_row("Brightness", -100, 100, 0)
        self._sl_contrast   = slider_row("Contrast",   -100, 100, 0)
        self._sl_sharpness  = slider_row("Sharpness",  -100, 100, 0)
        self._sl_blur       = slider_row("Blur",           0,  30, 0)

        layout.addWidget(self._sep())

        lbl_opt = QLabel("Options")
        lbl_opt.setObjectName("panelTitle")
        layout.addWidget(lbl_opt)

        self._chk_gray   = QCheckBox("Grayscale")
        self._chk_invert = QCheckBox("Invert colors")
        self._chk_gray.stateChanged.connect(lambda _: self._apply_filters())
        self._chk_invert.stateChanged.connect(lambda _: self._apply_filters())
        layout.addWidget(self._chk_gray)
        layout.addWidget(self._chk_invert)

        layout.addWidget(self._sep())

        lbl_spec = QLabel("Special Filters")
        lbl_spec.setObjectName("panelTitle")
        layout.addWidget(lbl_spec)

        for label, key in [
            ("Auto Contrast",          "auto"),
            ("Histogram Equalization", "equalize"),
            ("Edge Detection",         "edge"),
            ("Sharpen",                "sharpen"),
        ]:
            btn = QPushButton(label)
            btn.clicked.connect(lambda _, k=key: self._apply_special(k))
            layout.addWidget(btn)

        layout.addWidget(self._sep())
        reset_btn = QPushButton("\u21ba  Reset Filters")
        reset_btn.clicked.connect(self._reset_filters)
        layout.addWidget(reset_btn)
        layout.addStretch()
        return w

    def _build_tooth_context(self) -> QGroupBox:
        grp = QGroupBox("Tooth Context")
        grp.setStyleSheet("""
            QGroupBox {
                font-weight: 700; color: #14532D;
                border: 1px solid #86EFAC; border-radius: 6px;
                margin-top: 8px; padding: 8px 6px 6px 6px;
                background: #ECFDF5;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 8px; padding: 0 4px;
                background: #ECFDF5;
            }
        """)
        grid = QGridLayout(grp)
        grid.setSpacing(6)
        grid.setContentsMargins(6, 4, 6, 6)

        def combo(items):
            cb = QComboBox()
            cb.addItems(items)
            cb.setStyleSheet("font-size: 12px; min-height: 24px;")
            return cb

        self._cb_tooth = combo(["I1", "I2", "C", "Pm3", "Pm4", "M1", "M2", "M3"])
        self._cb_side  = combo(["Right", "Left"])
        self._cb_jaw   = combo(["Upper", "Lower"])

        grid.addWidget(QLabel("Tooth:"), 0, 0)
        grid.addWidget(self._cb_tooth,   0, 1)
        grid.addWidget(QLabel("Side:"),  0, 2)
        grid.addWidget(self._cb_side,    0, 3)
        grid.addWidget(QLabel("Jaw:"),   1, 0)
        grid.addWidget(self._cb_jaw,     1, 1)

        # Puech legend with colour dots
        legend = QLabel(
            "<small><b>Puech:</b> "
            "<span style='color:#3B82F6'>■</span>1=Horiz &nbsp;"
            "<span style='color:#EF4444'>■</span>2=Vert &nbsp;"
            "<span style='color:#22C55E'>■</span>3=Mesial &nbsp;"
            "<span style='color:#F97316'>■</span>4=Distal</small>"
        )
        legend.setStyleSheet("color: #4B7A5A;")
        legend.setTextFormat(Qt.TextFormat.RichText)
        grid.addWidget(legend, 1, 2, 1, 2)

        # Connect
        for cb in (self._cb_tooth, self._cb_side, self._cb_jaw):
            cb.currentTextChanged.connect(self._on_tooth_context_changed)

        return grp

    def _on_tooth_context_changed(self):
        jaw   = self._cb_jaw.currentText()
        side  = self._cb_side.currentText()
        tooth = self._cb_tooth.currentText()
        self.canvas.set_tooth_context(jaw, side, tooth)
        self._refresh_table()

    @staticmethod
    def _sep() -> QFrame:
        f = QFrame()
        f.setFrameShape(QFrame.Shape.HLine)
        f.setStyleSheet("background: #BBF7D0; max-height: 1px; border: none;")
        return f

    # ── Filter methods ────────────────────────────────────────────────────

    def _apply_filters(self):
        self._special_filter = None
        self.canvas.apply_filters(
            brightness=self._sl_brightness.value(),
            contrast=self._sl_contrast.value(),
            sharpness=self._sl_sharpness.value(),
            blur=self._sl_blur.value(),
            grayscale=self._chk_gray.isChecked(),
            invert=self._chk_invert.isChecked(),
        )

    def _apply_special(self, key: str):
        self._special_filter = key
        self.canvas.apply_filters(
            brightness=self._sl_brightness.value(),
            contrast=self._sl_contrast.value(),
            sharpness=self._sl_sharpness.value(),
            blur=self._sl_blur.value(),
            grayscale=self._chk_gray.isChecked(),
            invert=self._chk_invert.isChecked(),
            special=key,
        )

    def _reset_filters(self):
        self._special_filter = None
        for sl in (self._sl_brightness, self._sl_contrast,
                   self._sl_sharpness, self._sl_blur):
            sl.blockSignals(True)
            sl.setValue(0)
            sl.blockSignals(False)
        for chk in (self._chk_gray, self._chk_invert):
            chk.blockSignals(True)
            chk.setChecked(False)
            chk.blockSignals(False)
        self.canvas.reset_filters()

    # ── Menu bar ──────────────────────────────────────────────────────────

    def _build_menu(self):
        mb = self.menuBar()

        fm = mb.addMenu("File")
        self._add_action(fm, "\U0001f4c2  Open Image\u2026",  "Ctrl+O", self._open_image)
        self._add_action(fm, "\U0001f4c1  Open Folder\u2026", "Ctrl+Shift+O", self._open_folder)
        fm.addSeparator()
        self._add_action(fm, "\U0001f4be  Export\u2026",      "Ctrl+E", self._export)
        self._add_action(fm, "\U0001f5bc\ufe0f  Save Image\u2026", "Ctrl+Shift+S", self._save_image)
        fm.addSeparator()
        self._add_action(fm, "Quit", "Ctrl+Q", self.close)

        vm = mb.addMenu("View")
        self._add_action(vm, "Fit to Window", "Ctrl+0",       self.canvas.fit_view)
        self._add_action(vm, "Zoom In",       "Ctrl+=",       self.canvas.zoom_in)
        self._add_action(vm, "Zoom Out",      "Ctrl+-",       self.canvas.zoom_out)
        self._add_action(vm, "Reset Zoom",    "Ctrl+Shift+0", self.canvas.zoom_reset)

        em = mb.addMenu("Edit")
        self._add_action(em, "↻  Rotate 90° CW",          "Ctrl+Shift+R", lambda: self._rotate(90))
        self._add_action(em, "↺  Rotate 90° CCW",         None,           lambda: self._rotate(-90))
        self._add_action(em, "↕  Rotate 180°",            None,           lambda: self._rotate(180))
        em.addSeparator()
        self._add_action(em, "↔  Flip Horizontal",        None,           lambda: self._flip("h"))
        self._add_action(em, "↕  Flip Vertical",          None,           lambda: self._flip("v"))
        em.addSeparator()
        self._add_action(em, "📏  Scale Bar…",            None,           self._show_scale_bar)
        self._add_action(em, "✕  Remove Scale Bar",       None,           self.canvas.remove_scale_bar)

        am = mb.addMenu("Analysis")
        self._add_action(am, "Clear All Measurements", None, self._clear_all)
        am.addSeparator()
        # Striation type panel toggle
        self._act_show_stype = QAction("Show Striation Type Panel", self)
        self._act_show_stype.setCheckable(True)
        self._act_show_stype.setChecked(False)
        self._act_show_stype.toggled.connect(self._stype_panel.setVisible)
        am.addAction(self._act_show_stype)
        self._add_action(am, "\U0001f3f7  Rename / Add Types…", None, self._rename_stype)
        am.addSeparator()
        self._add_action(am, "\U0001f4cd  2D Landmarking & Procrustes…", "Ctrl+L",
                         self._open_landmarking)

        hm = mb.addMenu("Help")
        self._add_action(hm, "About Wearmap", None, lambda: AboutDialog(self).exec())

    def _add_action(self, menu, text, shortcut, slot):
        act = QAction(text, self)
        if shortcut:
            act.setShortcut(shortcut)
        act.triggered.connect(slot)
        menu.addAction(act)

    # ── Tool activation (handles mutual exclusivity) ──────────────────

    def _activate_tool(self, tool: str, act=None):
        for a in self._all_tool_acts:
            a.setChecked(False)
        if act is not None:
            act.setChecked(True)
        if act in self._measure_acts:
            self._measure_btn.setDefaultAction(act)
        self.canvas.set_tool(tool)

    # ── Toolbar ────────────────────────────────────────────────────

    def _build_toolbar(self):
        tb = self.addToolBar("Tools")
        tb.setIconSize(QSize(24, 24))
        tb.setMovable(False)

        self._all_tool_acts: list = []
        self._measure_acts:  list = []

        open_act = QAction("📂  Open", self)
        open_act.setToolTip("Open image (Ctrl+O)")
        open_act.triggered.connect(self._open_image)
        tb.addAction(open_act)
        tb.addSeparator()

        self._prev_btn = QAction("◀  Prev", self)
        self._prev_btn.setToolTip("Previous image in folder (Left arrow)")
        self._prev_btn.setEnabled(False)
        self._prev_btn.triggered.connect(lambda: self._navigate(-1))
        tb.addAction(self._prev_btn)

        self._nav_lbl = QLabel("  —  ")
        self._nav_lbl.setStyleSheet("color: #166534; padding: 0 4px; font-size: 12px;")
        tb.addWidget(self._nav_lbl)

        self._next_btn = QAction("Next  ▶", self)
        self._next_btn.setToolTip("Next image in folder (Right arrow)")
        self._next_btn.setEnabled(False)
        self._next_btn.triggered.connect(lambda: self._navigate(1))
        tb.addAction(self._next_btn)
        tb.addSeparator()

        # Pan
        self._act_select = QAction("☞  Pan", self)
        self._act_select.setToolTip("Pan / drag the image  [S]")
        self._act_select.setCheckable(True)
        self._act_select.setChecked(True)
        self._act_select.triggered.connect(
            lambda: self._activate_tool(Tool.SELECT, self._act_select)
        )
        self._all_tool_acts.append(self._act_select)
        tb.addAction(self._act_select)

        # Measure dropdown
        self._measure_btn = QToolButton()
        self._measure_btn.setPopupMode(
            QToolButton.ToolButtonPopupMode.MenuButtonPopup
        )
        self._measure_btn.setToolTip(
            "Measurement tools\nClick = use last tool · Arrow = switch tool"
        )
        measure_menu = QMenu(self._measure_btn)
        self._measure_btn.setMenu(measure_menu)

        for _icon, _label, _tip, _tool in [
            ("↔", "Distance", "Measure distance  [D]",                           Tool.DISTANCE),
            ("⬡", "Area",     "Measure area  [A] — right-click to close", Tool.AREA),
            ("∠", "Angle",    "Measure angle  [G]",                              Tool.ANGLE),
            ("#",      "Count",    "Count objects  [C] — right-click to finish", Tool.COUNT),
            ("⋅", "Pits",     "Mark pits  [P] — right-click to finish",   Tool.PITS),
        ]:
            _act = QAction(f"{_icon}  {_label}", self)
            _act.setToolTip(_tip)
            _act.setCheckable(True)
            _act.triggered.connect(
                lambda _, t=_tool, a=_act: self._activate_tool(t, a)
            )
            self._all_tool_acts.append(_act)
            self._measure_acts.append(_act)
            measure_menu.addAction(_act)

        (self._act_dist, self._act_area, self._act_angle,
         self._act_count, self._act_pits) = self._measure_acts
        self._measure_btn.setDefaultAction(self._act_dist)
        tb.addWidget(self._measure_btn)

        # Calibrate
        self._act_cal = QAction("⊞  Calibrate", self)
        self._act_cal.setToolTip("Set scale calibration  [K]")
        self._act_cal.setCheckable(True)
        self._act_cal.triggered.connect(
            lambda: self._activate_tool(Tool.CALIBRATE, self._act_cal)
        )
        self._all_tool_acts.append(self._act_cal)
        tb.addAction(self._act_cal)
        tb.addSeparator()

        for icon, tip, fn in [
            ("🔍+",      "Zoom in (Ctrl+=)",          self.canvas.zoom_in),
            ("🔍−", "Zoom out (Ctrl+-)",         self.canvas.zoom_out),
            ("⊟  Fit",      "Fit to window (Ctrl+0)",    self.canvas.fit_view),
        ]:
            a = QAction(icon, self)
            a.setToolTip(tip)
            a.triggered.connect(fn)
            tb.addAction(a)

        self._act_labels = QAction("🏷️  Labels", self)
        self._act_labels.setToolTip("Show / hide measurement labels  [L]")
        self._act_labels.setCheckable(True)
        self._act_labels.setChecked(True)
        self._act_labels.toggled.connect(self.canvas.set_labels_visible)
        tb.addAction(self._act_labels)

        self._act_puech_colors = QAction("🎨  Puech Colors", self)
        self._act_puech_colors.setToolTip(
            "Color striations by Puech class\n"
            "Blue=P1 Horiz · Red=P2 Vert · Green=P3 Mesial · Orange=P4 Distal"
        )
        self._act_puech_colors.setCheckable(True)
        self._act_puech_colors.setChecked(False)
        self._act_puech_colors.toggled.connect(self.canvas.set_puech_colors)
        tb.addAction(self._act_puech_colors)
        tb.addSeparator()

        exp_act = QAction("💾  Export", self)
        exp_act.setToolTip("Export measurements (Ctrl+E)")
        exp_act.triggered.connect(self._export)
        tb.addAction(exp_act)

        img_act = QAction("🖼️  Save Image", self)
        img_act.setToolTip("Save annotated image (Ctrl+Shift+S)")
        img_act.triggered.connect(self._save_image)
        tb.addAction(img_act)

    # ── Keyboard shortcuts ────────────────────────────────────────────────

    def _build_shortcuts(self):
        # Tool shortcuts
        for key, act, tool in [
            ("S", self._act_select, Tool.SELECT),
            ("D", self._act_dist,   Tool.DISTANCE),
            ("A", self._act_area,   Tool.AREA),
            ("G", self._act_angle,  Tool.ANGLE),
            ("C", self._act_count,  Tool.COUNT),
            ("P", self._act_pits,   Tool.PITS),
            ("K", self._act_cal,    Tool.CALIBRATE),
        ]:
            sc = QShortcut(QKeySequence(key), self)
            sc.activated.connect(
                lambda a=act, t=tool: self._activate_tool(t, a)
            )

        # Striation type shortcuts: 1–9 assign the type to selected rows
        for n in range(1, 10):
            sc = QShortcut(QKeySequence(str(n)), self)
            sc.activated.connect(lambda k=n: self._stype_shortcut(k))

        QShortcut(QKeySequence("Delete"), self).activated.connect(self._delete_selected)
        QShortcut(QKeySequence("Left"),   self).activated.connect(lambda: self._navigate(-1))
        QShortcut(QKeySequence("Right"),  self).activated.connect(lambda: self._navigate(1))
        QShortcut(QKeySequence("L"),      self).activated.connect(
            lambda: self._act_labels.setChecked(not self._act_labels.isChecked())
        )

    # ── Status bar ────────────────────────────────────────────────────────

    def _build_statusbar(self):
        sb = self.statusBar()
        style = "padding: 0 10px; color: #166534;"
        self._pos_lbl  = QLabel("x=\u2014  y=\u2014")
        self._pos_lbl.setStyleSheet(style)
        self._cal_lbl  = QLabel("Calibration: px")
        self._cal_lbl.setStyleSheet(style)
        self._hint_lbl = QLabel("Open an image to start")
        self._hint_lbl.setStyleSheet("color: #4ADE80; font-style: italic; padding: 0 10px;")
        sb.addWidget(self._pos_lbl)
        sb.addWidget(QLabel("|"))
        sb.addWidget(self._cal_lbl)
        sb.addWidget(QLabel("|"))
        sb.addWidget(self._hint_lbl)

    # ── Signal handlers ───────────────────────────────────────────────────

    def _on_measurement_added(self, m):
        self._refresh_table()
        # Auto-target the just-drawn measurement so 1/2/3 immediately applies to it
        self._stype_target_mids = {m.id}

    def _on_table_selection_changed(self):
        self._stype_target_mids = set()
        for item in self.table.selectedItems():
            if item.column() == 0:
                try:
                    self._stype_target_mids.add(int(item.text()))
                except ValueError:
                    pass

    def _on_canvas_measurement_clicked(self, mid: int):
        self._stype_target_mids = {mid}
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and int(item.text()) == mid:
                self.table.selectRow(row)
                self.table.scrollToItem(item)
                break

    def _update_position(self, x: float, y: float):
        self._pos_lbl.setText(f"x={x:.1f}  y={y:.1f}")

    def _on_calibration_line(self, pixel_length: float):
        dlg = CalibrationDialog(pixel_length, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            cal = dlg.get_calibration()
            self.canvas.update_calibration(cal)
            self._refresh_table()
            unit = cal.unit if cal.is_calibrated else "px"
            self.cal_label.setText(
                f"Scale: 1 px \u2192 {cal.scale:.6g} {unit}\nUnit: <b>{unit}</b>"
            )
            self._cal_lbl.setText(f"Calibration: {unit}")
        self._act_select.setChecked(True)
        self.canvas.set_tool(Tool.SELECT)

    # ── Image loading ─────────────────────────────────────────────────────

    def _open_image(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Image", "",
            "Images (*.png *.jpg *.jpeg *.tif *.tiff *.bmp *.gif *.webp *.plu *.plux);;All Files (*)",
        )
        if path:
            self._load(path, scan_folder=True)

    def _open_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Open Folder")
        if folder:
            files = self._scan_folder(folder)
            if not files:
                QMessageBox.information(self, "Open Folder", "No image files found in that folder.")
                return
            self._load(files[0], scan_folder=False, files=files)

    def _scan_folder(self, folder: str) -> list[str]:
        files = sorted(
            os.path.join(folder, f) for f in os.listdir(folder)
            if os.path.splitext(f)[1].lower() in IMAGE_EXTS
        )
        return files

    def _load(self, path: str, scan_folder: bool = True, files: list[str] | None = None):
        # Save current image state before switching
        if self._current_path:
            self._image_states[self._current_path] = self.canvas.get_state()

        if not self.canvas.load_image(path):
            QMessageBox.critical(self, "Error", f"Could not load image:\n{path}")
            return

        self._current_path = path
        self._reset_filters()

        # Build file list for navigation
        if files is not None:
            self._image_files = files
        elif scan_folder:
            self._image_files = self._scan_folder(os.path.dirname(path))

        if self._image_files and path in self._image_files:
            self._image_idx = self._image_files.index(path)
        else:
            self._image_idx = -1

        # Restore previous measurements + tooth context if we've visited this image
        if path in self._image_states:
            self.canvas.restore_state(self._image_states[path])
            cal = self.canvas.calibration
            unit = cal.unit if cal.is_calibrated else "px"
            self.cal_label.setText(f"Scale: 1 px \u2192 {cal.scale:.6g} {unit}\nUnit: <b>{unit}</b>")
            self._cal_lbl.setText(f"Calibration: {unit}")
            # Sync dropdowns to restored state (block signals to avoid loop)
            for cb, key in [(self._cb_jaw,   "jaw"),
                            (self._cb_side,  "side"),
                            (self._cb_tooth, "tooth")]:
                cb.blockSignals(True)
                cb.setCurrentText(self._image_states[path].get(key, cb.currentText()))
                cb.blockSignals(False)

        self._refresh_table()
        self._update_nav_ui()

        w, h = self.canvas.image_size()
        try:
            size_kb = os.path.getsize(path) // 1024
        except OSError:
            size_kb = 0
        self.info_label.setText(
            f"<b>{os.path.basename(path)}</b><br>"
            f"Dimensions: {w} \u00d7 {h} px<br>"
            f"File size: {size_kb} KB"
        )
        self.setWindowTitle(f"Wearmap \u2014 {os.path.basename(path)}")
        self._hint_lbl.setText("Choose a tool from the toolbar to start measuring")

    # ── Navigation ────────────────────────────────────────────────────────

    def _navigate(self, delta: int):
        if not self._image_files or self._image_idx < 0:
            return
        new_idx = self._image_idx + delta
        if not (0 <= new_idx < len(self._image_files)):
            return
        self._load(self._image_files[new_idx], scan_folder=False)

    def _update_nav_ui(self):
        n = len(self._image_files)
        if n > 0 and self._image_idx >= 0:
            self._nav_lbl.setText(f"  {self._image_idx + 1} / {n}  ")
            self._prev_btn.setEnabled(self._image_idx > 0)
            self._next_btn.setEnabled(self._image_idx < n - 1)
        else:
            self._nav_lbl.setText("  \u2014  ")
            self._prev_btn.setEnabled(False)
            self._next_btn.setEnabled(False)

    # ── Export ────────────────────────────────────────────────────────────

    def _all_measurements(self) -> list[tuple]:
        """Returns [(image_name, measurement, cal, jaw, side, tooth), ...] across all images."""
        result = []
        for p, state in self._image_states.items():
            name = os.path.basename(p)
            jaw, side, tooth = state.get("jaw","Upper"), state.get("side","Right"), state.get("tooth","")
            for m in state["measurements"]:
                result.append((name, m, state["calibration"], jaw, side, tooth))
        if self._current_path and self._current_path not in self._image_states:
            name = os.path.basename(self._current_path)
            for m in self.canvas.measurements:
                result.append((name, m, self.canvas.calibration,
                               self.canvas.jaw, self.canvas.side, self.canvas.tooth))
        return result

    def _show_scale_bar(self):
        dlg = ScaleBarDialog(self.canvas.calibration, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.canvas.add_scale_bar(dlg.value, dlg.unit)

    def _rotate(self, degrees: int):
        if not self.canvas.pixmap_item:
            return
        self.canvas.rotate_image(degrees)
        self._refresh_table()

    def _flip(self, axis: str):
        if self.canvas._original_pil is None:
            return
        from PIL import ImageOps
        if axis == "h":
            self.canvas._original_pil = ImageOps.mirror(self.canvas._original_pil)
        else:
            self.canvas._original_pil = ImageOps.flip(self.canvas._original_pil)
        self.canvas.rotate_image(0)   # reload with 0° (just redraws)

    def _save_image(self):
        if not self.canvas.pixmap_item:
            QMessageBox.information(self, "Save Image", "No image loaded.")
            return
        default = os.path.splitext(os.path.basename(self._current_path))[0] + "_annotated" if self._current_path else "annotated"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Annotated Image", default,
            "PNG (*.png);;TIFF (*.tiff);;JPEG (*.jpg)",
        )
        if not path:
            return
        ok = self.canvas.save_annotated_image(path)
        if ok:
            QMessageBox.information(self, "Save Image", f"Saved:\n{path}")
        else:
            QMessageBox.critical(self, "Save Image", "Could not save the image.")

    def _export(self):
        if self._current_path:
            self._image_states[self._current_path] = self.canvas.get_state()

        all_rows = self._all_measurements()
        if not all_rows:
            QMessageBox.information(self, "Export", "No measurements to export yet.")
            self._image_states.pop(self._current_path, None)
            return

        path, sel = QFileDialog.getSaveFileName(
            self, "Export Measurements", "measurements",
            "Excel (*.xlsx);;CSV (*.csv);;Tab-separated (*.txt)",
        )
        if not path:
            self._image_states.pop(self._current_path, None)
            return

        try:
            detail_hdrs = ["Image", "ID", "Type", "Tooth", "Jaw", "Side",
                           "Unit", "Value", "Slope (°)", "Puech", "Striation Type", "Points (px)"]
            detail_rows = []
            for img_name, m, cal, jaw, side, tooth in all_rows:
                row_data = list(self.canvas._row_data_with_cal(m, cal, jaw, side, tooth))
                # Replace numeric striation_type with human name (index 9 = stype col)
                stype_idx = 9  # 0-based in row_data: id,type,tooth,jaw,side,unit,val,slope,puech,stype,pts
                if len(row_data) > stype_idx and row_data[stype_idx] not in ("-", "0", 0):
                    try:
                        sn = int(row_data[stype_idx])
                        row_data[stype_idx] = (
                            f"{sn}: {self._stype_names.get(sn, f'Type {sn}')}"
                            if sn else "-"
                        )
                    except (ValueError, TypeError):
                        pass
                detail_rows.append([img_name] + row_data)

            summ_hdrs = ["Image", "Puech", "Category", "Count",
                         "Mean Distance", "Std Dev", "Unit"]
            summ_rows = self._build_summary_rows(all_rows)

            if "xlsx" in path or "xlsx" in sel:
                if not path.endswith(".xlsx"):
                    path += ".xlsx"
                self._write_xlsx_two_sheets(path, detail_hdrs, detail_rows,
                                            summ_hdrs, summ_rows)
                saved = path
            else:
                is_txt = "txt" in path or "txt" in sel
                ext    = ".txt" if is_txt else ".csv"
                base   = path if path.endswith(ext) else path + ext
                summ_path = base.replace(ext, f"_summary{ext}")
                if is_txt:
                    self._write_txt(base, detail_hdrs, detail_rows)
                    self._write_txt(summ_path, summ_hdrs, summ_rows)
                else:
                    self._write_csv(base, detail_hdrs, detail_rows)
                    self._write_csv(summ_path, summ_hdrs, summ_rows)
                saved = f"{base}\n{summ_path}"

            QMessageBox.information(self, "Export", f"Saved successfully:\n{saved}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    @staticmethod
    def _build_summary_rows(all_rows: list) -> list:
        import math
        from collections import defaultdict
        LABELS = {1: "Horizontal", 2: "Vertical", 3: "Mesial", 4: "Distal"}
        # img -> puech -> (distances, unit)
        data: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
        units: dict[str, str] = {}
        for img_name, m, cal, jaw, side, tooth in all_rows:
            if m.mtype != MeasurementType.DISTANCE:
                continue
            p = m.puech(jaw, side)
            if p == 0:
                continue
            data[img_name][p].append(cal.to_real(m.pixel_value()))
            units[img_name] = cal.unit_str()
        def _stats(vals):
            n = len(vals)
            if n == 0:
                return n, "-", "-"
            if n == 1:
                return n, f"{vals[0]:.4f}", "0.0000"
            mean = sum(vals) / n
            std  = math.sqrt(sum((v - mean) ** 2 for v in vals) / (n - 1))
            return n, f"{mean:.4f}", f"{std:.4f}"

        rows = []
        for img_name in sorted(data):
            for p in [1, 2, 3, 4]:
                vals = data[img_name].get(p, [])
                n, mean_s, std_s = _stats(vals)
                rows.append([img_name, p, LABELS[p], n, mean_s, std_s,
                             units.get(img_name, "px")])
            # Total row across all Puech categories for this image
            all_vals = [v for p_vals in data[img_name].values() for v in p_vals]
            n_tot, mean_tot, std_tot = _stats(all_vals)
            rows.append([img_name, "TOTAL", "", n_tot, mean_tot, std_tot,
                         units.get(img_name, "px")])
        return rows

    @staticmethod
    def _write_xlsx_two_sheets(path, detail_hdrs, detail_rows, summ_hdrs, summ_rows):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        def fill_sheet(ws, headers, rows, hdr_color, alt_color):
            hdr_fill  = PatternFill("solid", fgColor=hdr_color)
            hdr_font  = Font(bold=True, color="14532D")
            tot_fill  = PatternFill("solid", fgColor="BBF7D0")
            tot_font  = Font(bold=True, color="14532D")
            alt_fill  = PatternFill("solid", fgColor=alt_color)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(rows, 2):
                is_total = len(row) > 1 and row[1] == "TOTAL"
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    if is_total:
                        cell.fill = tot_fill
                        cell.font = tot_font
                    elif ri % 2 == 0:
                        cell.fill = alt_fill
            for col_cells in ws.columns:
                width = max(len(str(c.value or "")) for c in col_cells) + 4
                ws.column_dimensions[col_cells[0].column_letter].width = min(width, 50)

        ws1 = wb.active
        ws1.title = "Detail"
        fill_sheet(ws1, detail_hdrs, detail_rows, "86EFAC", "ECFDF5")

        ws2 = wb.create_sheet("Summary")
        fill_sheet(ws2, summ_hdrs, summ_rows, "4ADE80", "F0FFF4")

        wb.save(path)

    @staticmethod
    def _write_csv(path, headers, rows):
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)

    @staticmethod
    def _write_txt(path, headers, rows):
        with open(path, "w", encoding="utf-8") as f:
            f.write("\t".join(headers) + "\n")
            for r in rows:
                f.write("\t".join(str(x) for x in r) + "\n")


    # ── Table management ──────────────────────────────────────────────────

    def _refresh_table(self):
        self.table.setRowCount(0)
        jaw  = self.canvas.jaw
        side = self.canvas.side
        for m in self.canvas.measurements:
            row = self.table.rowCount()
            self.table.insertRow(row)
            # When Puech colors are active, use the Puech colour for Distance rows
            if self.canvas.use_puech_colors and m.mtype == MeasurementType.DISTANCE:
                from canvas import PUECH_COLORS
                p = m.puech(jaw, side)
                color = QColor(PUECH_COLORS.get(p, m.color))
            else:
                color = QColor(m.color)
            slope_txt = m.slope_display() or "\u2014"
            puech_n   = m.puech(jaw, side)
            puech_txt = str(puech_n) if puech_n else "\u2014"
            stype_n   = m.striation_type
            stype_txt = (
                self._stype_names.get(stype_n, f"Type {stype_n}")
                if stype_n else "\u2014"
            )
            for col, text in enumerate([
                str(m.id),
                m.mtype.value,
                m.display_value(self.canvas.calibration),
                slope_txt,
                puech_txt,
                stype_txt,
                m.label,
            ]):
                item = QTableWidgetItem(text)
                item.setForeground(color)
                self.table.setItem(row, col, item)

    # ── Striation type helpers ────────────────────────────────────────────

    def _rebuild_stype_combo(self):
        self._stype_combo.blockSignals(True)
        self._stype_combo.clear()
        self._stype_combo.addItem("—  (not set)", 0)
        for k in sorted(self._stype_names.keys()):
            self._stype_combo.addItem(f"{k}: {self._stype_names[k]}", k)
        self._stype_combo.blockSignals(False)

    def _on_stype_combo_changed(self, index=0):
        stype = self._stype_combo.currentData()
        if stype is None:
            return
        self._assign_stype_value(stype)

    def _apply_striation_type(self):
        stype = self._stype_combo.currentData()
        self._assign_stype_value(stype)

    def _stype_shortcut(self, n: int):
        if n not in self._stype_names:
            return
        if not self._assign_stype_value(n):
            return
        for i in range(self._stype_combo.count()):
            if self._stype_combo.itemData(i) == n:
                self._stype_combo.blockSignals(True)
                self._stype_combo.setCurrentIndex(i)
                self._stype_combo.blockSignals(False)
                break

    def _assign_stype_value(self, stype: int) -> bool:
        # Fallback: if nothing selected, target the last-drawn measurement
        targets = set(self._stype_target_mids)
        if not targets and self.canvas.measurements:
            targets = {self.canvas.measurements[-1].id}
        if not targets:
            return False
        for m in self.canvas.measurements:
            if m.id in targets:
                m.striation_type = stype
        self._refresh_table()
        # Re-select the same rows by measurement ID (use local var; signal
        # cascade from _refresh_table may have cleared self._stype_target_mids)
        from PyQt6.QtCore import QItemSelectionModel
        sel = self.table.selectionModel()
        sel.blockSignals(True)
        sel.clearSelection()
        flags = (QItemSelectionModel.SelectionFlag.Select |
                 QItemSelectionModel.SelectionFlag.Rows)
        for row in range(self.table.rowCount()):
            id_item = self.table.item(row, 0)
            if id_item and int(id_item.text()) in targets:
                sel.select(self.table.model().index(row, 0), flags)
        sel.blockSignals(False)
        self._stype_target_mids = targets
        return True

    def _rename_stype(self):
        dlg = StriationTypeNamesDialog(self._stype_names, self)
        if dlg.exec():
            self._stype_names = dlg.get_names()
            self._rebuild_stype_combo()
            self._refresh_table()

    def _open_landmarking(self):
        if self._landmarking_dlg is None or not self._landmarking_dlg.isVisible():
            self._landmarking_dlg = LandmarkingDialog(self)
        self._landmarking_dlg.show()
        self._landmarking_dlg.raise_()

    def _delete_selected(self):
        rows = sorted({i.row() for i in self.table.selectedItems()}, reverse=True)
        for row in rows:
            item = self.table.item(row, 0)
            if item:
                self.canvas.delete_by_id(int(item.text()))
        self._refresh_table()

    def _clear_all(self):
        if not self.canvas.measurements:
            return
        reply = QMessageBox.question(
            self, "Clear All", "Delete all measurements for this image?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.canvas.clear_all()
            self.table.setRowCount(0)

    # ── Drag & drop ───────────────────────────────────────────────────────

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            if os.path.isfile(path):
                self._load(path, scan_folder=True)
