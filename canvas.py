import io
import math
import os
from typing import List, Optional, Tuple

from PyQt6.QtCore import Qt, QPointF, QRectF, pyqtSignal
from PyQt6.QtGui import (
    QBrush, QColor, QFont, QFontMetricsF, QImage, QPainter, QPen, QPixmap,
    QPolygonF, QImageReader,
)
from PyQt6.QtWidgets import QGraphicsItem, QGraphicsView, QGraphicsScene

from measurements import Calibration, Measurement, MeasurementType, compute_puech

PUECH_COLORS = {
    1: "#3B82F6",  # Horizontal - blue
    2: "#EF4444",  # Vertical   - red
    3: "#22C55E",  # Mesial     - green
    4: "#F97316",  # Distal     - orange
}


class ScaleBarItem(QGraphicsItem):
    """Visual scale bar rendered in scene (image) coordinates."""
    _PAD = 8
    _BAR_H = 7
    _TICK_EXTRA = 4   # extra length above/below bar on each tick

    def __init__(self, bar_px: float, label: str):
        super().__init__()
        self._bar_px = bar_px
        self._label  = label
        self._font   = QFont("Segoe UI", 10)
        self._font.setBold(True)
        fm = QFontMetricsF(self._font)
        self._text_w = fm.horizontalAdvance(label)
        self._text_h = fm.ascent()

    def boundingRect(self) -> QRectF:
        w = max(self._bar_px, self._text_w) + 2 * self._PAD
        h = self._text_h + 6 + self._BAR_H + 2 * self._TICK_EXTRA + 2 * self._PAD
        return QRectF(-self._PAD,
                      -(self._text_h + 6 + self._TICK_EXTRA + self._PAD),
                      w, h)

    def paint(self, painter: QPainter, option, widget=None):
        painter.save()
        # Background
        painter.fillRect(self.boundingRect(), QColor(0, 0, 0, 155))
        # Tick marks
        t_pen = QPen(QColor(255, 255, 255), 2)
        painter.setPen(t_pen)
        for x in (0.0, self._bar_px):
            painter.drawLine(QPointF(x, -self._TICK_EXTRA),
                             QPointF(x, self._BAR_H + self._TICK_EXTRA))
        # Horizontal bar
        b_pen = QPen(QColor(255, 255, 255), self._BAR_H)
        painter.setPen(b_pen)
        painter.drawLine(QPointF(0, self._BAR_H / 2),
                         QPointF(self._bar_px, self._BAR_H / 2))
        # Label centred above bar
        painter.setFont(self._font)
        painter.setPen(QPen(QColor(255, 255, 255)))
        tx = (self._bar_px - self._text_w) / 2
        painter.drawText(QPointF(tx, -(self._TICK_EXTRA + 3)), self._label)
        painter.restore()


class Tool:
    SELECT = "select"
    DISTANCE = "distance"
    AREA = "area"
    ANGLE = "angle"
    COUNT = "count"
    PITS = "pits"
    CALIBRATE = "calibrate"


class MeasurementItem(QGraphicsItem):
    def __init__(self, measurement: Measurement, calibration: Calibration,
                 jaw: str = "Upper", side: str = "Right", show_labels: bool = True,
                 use_puech_colors: bool = False):
        super().__init__()
        self.measurement      = measurement
        self.calibration      = calibration
        self.jaw              = jaw
        self.side             = side
        self.show_labels      = show_labels
        self.use_puech_colors = use_puech_colors
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)

    def boundingRect(self) -> QRectF:
        pts = self.measurement.points
        if not pts:
            return QRectF(0, 0, 1, 1)
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        margin = 70
        return QRectF(
            min(xs) - margin,
            min(ys) - margin,
            max(xs) - min(xs) + 2 * margin,
            max(ys) - min(ys) + 2 * margin,
        )

    def paint(self, painter: QPainter, option, widget=None):
        if not self.measurement.visible or not self.measurement.points:
            return
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        color = QColor(self.measurement.color)
        mtype = self.measurement.mtype
        if self.use_puech_colors and mtype == MeasurementType.DISTANCE:
            p = self.measurement.puech(self.jaw, self.side)
            if p in PUECH_COLORS:
                color = QColor(PUECH_COLORS[p])
        pen = QPen(color, 2.0)
        pen.setCosmetic(True)
        pts = self.measurement.points
        if mtype == MeasurementType.DISTANCE:
            self._paint_distance(painter, pts, color, pen)
        elif mtype == MeasurementType.AREA:
            self._paint_area(painter, pts, color, pen)
        elif mtype == MeasurementType.ANGLE:
            self._paint_angle(painter, pts, color, pen)
        elif mtype == MeasurementType.COUNT:
            self._paint_count(painter, pts, color)
        elif mtype == MeasurementType.PITS:
            self._paint_pits(painter, pts, color)

    # ── helpers ──────────────────────────────────────────────────────────

    def _handle(self, painter: QPainter, x: float, y: float, color: QColor, r: float = 3.0):
        painter.save()
        p = QPen(color, 1.5)
        p.setCosmetic(True)
        painter.setPen(p)
        painter.setBrush(QBrush(Qt.GlobalColor.white))
        painter.drawEllipse(QRectF(x - r, y - r, 2 * r, 2 * r))
        painter.restore()

    def _label(self, painter: QPainter, cx: float, cy: float, text: str, color: QColor):
        painter.save()
        font = QFont("Segoe UI", 9)
        font.setBold(True)
        painter.setFont(font)
        fm = QFontMetricsF(font)
        tw = fm.horizontalAdvance(text)
        th = fm.height()
        pad = 5.0
        rx, ry = cx - tw / 2 - pad, cy - th - pad
        rw, rh = tw + 2 * pad, th + 2 * pad
        bg = QColor(255, 255, 255, 215)
        painter.setPen(QPen(color, 0.8))
        painter.setBrush(QBrush(bg))
        painter.drawRoundedRect(QRectF(rx, ry, rw, rh), 3, 3)
        painter.setPen(QPen(color))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawText(QRectF(rx, ry, rw, rh), Qt.AlignmentFlag.AlignCenter, text)
        painter.restore()

    # ── measurement painters ─────────────────────────────────────────────

    def _paint_distance(self, painter, pts, color, pen):
        if len(pts) < 2:
            return
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawLine(QPointF(*pts[0]), QPointF(*pts[1]))
        # tick marks at ends
        angle = math.atan2(pts[1][1] - pts[0][1], pts[1][0] - pts[0][0])
        perp = angle + math.pi / 2
        tick = 7
        for p in (pts[0], pts[1]):
            x1 = p[0] + tick * math.cos(perp)
            y1 = p[1] + tick * math.sin(perp)
            x2 = p[0] - tick * math.cos(perp)
            y2 = p[1] - tick * math.sin(perp)
            painter.drawLine(QPointF(x1, y1), QPointF(x2, y2))
        self._handle(painter, *pts[0], color)
        self._handle(painter, *pts[1], color)
        mx = (pts[0][0] + pts[1][0]) / 2
        my = (pts[0][1] + pts[1][1]) / 2
        dist_txt  = self.measurement.display_value(self.calibration)
        slope_txt = self.measurement.slope_display()
        puech_n   = self.measurement.puech(self.jaw, self.side)
        puech_txt = f"  P{puech_n}" if puech_n else ""
        full_txt  = f"{dist_txt}   {slope_txt}{puech_txt}" if slope_txt else dist_txt
        if self.show_labels:
            self._label(painter, mx, my - 8, full_txt, color)

    def _paint_area(self, painter, pts, color, pen):
        if len(pts) < 2:
            return
        painter.setPen(pen)
        fill = QColor(color)
        fill.setAlpha(45)
        painter.setBrush(QBrush(fill))
        painter.drawPolygon(QPolygonF([QPointF(*p) for p in pts]))
        for p in pts:
            self._handle(painter, *p, color)
        if len(pts) >= 3:
            cx = sum(p[0] for p in pts) / len(pts)
            cy = sum(p[1] for p in pts) / len(pts)
            if self.show_labels:
                self._label(painter, cx, cy, self.measurement.display_value(self.calibration), color)

    def _paint_angle(self, painter, pts, color, pen):
        if len(pts) < 3:
            return
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawLine(QPointF(*pts[0]), QPointF(*pts[1]))
        painter.drawLine(QPointF(*pts[1]), QPointF(*pts[2]))

        a1 = (pts[0][0] - pts[1][0], pts[0][1] - pts[1][1])
        a2 = (pts[2][0] - pts[1][0], pts[2][1] - pts[1][1])
        ang1 = math.degrees(math.atan2(-a1[1], a1[0]))
        ang2 = math.degrees(math.atan2(-a2[1], a2[0]))
        span = ang2 - ang1
        if span > 180:
            span -= 360
        elif span < -180:
            span += 360

        arc_r = 32
        arc_rect = QRectF(pts[1][0] - arc_r, pts[1][1] - arc_r, 2 * arc_r, 2 * arc_r)
        arc_pen = QPen(color, 1.5)
        arc_pen.setCosmetic(True)
        painter.setPen(arc_pen)
        painter.drawArc(arc_rect, int(ang1 * 16), int(span * 16))

        for p in pts:
            self._handle(painter, *p, color)

        mid_rad = math.radians(ang1 + span / 2)
        lx = pts[1][0] + 52 * math.cos(mid_rad)
        ly = pts[1][1] - 52 * math.sin(mid_rad)
        if self.show_labels:
            self._label(painter, lx, ly, self.measurement.display_value(self.calibration), color)

    def _paint_pits(self, painter, pts, color):
        r = 4.0
        for p in pts:
            cp = QPen(color, 1.2)
            cp.setCosmetic(True)
            painter.setPen(cp)
            fill = QColor(color)
            fill.setAlpha(220)
            painter.setBrush(QBrush(fill))
            painter.drawEllipse(QRectF(p[0] - r, p[1] - r, 2 * r, 2 * r))

    def _paint_count(self, painter, pts, color):
        r = 11.0
        for i, p in enumerate(pts):
            cp = QPen(color, 2.0)
            cp.setCosmetic(True)
            painter.setPen(cp)
            fill = QColor(color)
            fill.setAlpha(210)
            painter.setBrush(QBrush(fill))
            painter.drawEllipse(QRectF(p[0] - r, p[1] - r, 2 * r, 2 * r))
            font = QFont("Segoe UI", 8)
            font.setBold(True)
            painter.setFont(font)
            painter.setPen(QPen(Qt.GlobalColor.white))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawText(
                QRectF(p[0] - r, p[1] - r, 2 * r, 2 * r),
                Qt.AlignmentFlag.AlignCenter,
                str(i + 1),
            )


# ── Canvas ────────────────────────────────────────────────────────────────────

class ImageCanvas(QGraphicsView):
    measurement_added = pyqtSignal(object)
    measurement_clicked = pyqtSignal(int)   # emits measurement id
    calibration_line_drawn = pyqtSignal(float)
    position_changed = pyqtSignal(float, float)

    COLORS = [
        "#22C55E", "#0EA5E9", "#F59E0B", "#EF4444",
        "#8B5CF6", "#EC4899", "#14B8A6", "#F97316",
    ]

    def __init__(self):
        super().__init__()
        self._scene = QGraphicsScene(self)
        self.setScene(self._scene)
        self.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.setBackgroundBrush(QBrush(QColor("#1e2030")))
        self.setStyleSheet("border: none;")

        self.pixmap_item = None
        self.calibration = Calibration()
        self.measurements: List[Measurement] = []
        self.measurement_items: List[MeasurementItem] = []
        self._next_id = 1
        self._color_idx = 0

        self.current_tool = Tool.SELECT
        self._in_progress: List[Tuple[float, float]] = []
        self._previews = []

        self._active_count: Optional[Measurement] = None
        self._active_count_item: Optional[MeasurementItem] = None

        self._original_pil = None

        # Tooth context (set from main window)
        self.jaw              = "Upper"
        self.side             = "Right"
        self.tooth            = "I1"
        self.show_labels      = True
        self.use_puech_colors = False
        self._scale_bar_item: Optional[ScaleBarItem] = None

    # ── Scale bar ─────────────────────────────────────────────────────────

    def add_scale_bar(self, value: float, unit: str):
        """Draw a scale bar at bottom-right of the image. Requires calibration."""
        if not self.pixmap_item or not self.calibration.is_calibrated:
            return
        bar_px = value / self.calibration.scale
        label  = f"{value:g} {unit}"
        self.remove_scale_bar()
        self._scale_bar_item = ScaleBarItem(bar_px, label)
        self._scene.addItem(self._scale_bar_item)
        rect   = self.pixmap_item.boundingRect()
        margin = 24
        x = rect.right()  - bar_px - ScaleBarItem._PAD - margin
        y = rect.bottom() - ScaleBarItem._BAR_H - ScaleBarItem._TICK_EXTRA - margin
        self._scale_bar_item.setPos(x, y)

    def remove_scale_bar(self):
        if self._scale_bar_item is not None:
            self._scene.removeItem(self._scale_bar_item)
            self._scale_bar_item = None

    # ── Image rotation ────────────────────────────────────────────────────

    def rotate_image(self, degrees: int):
        """Rotate the current image by degrees (90, -90, 180). Clears measurements."""
        if self._original_pil is None:
            return
        from PIL import Image
        self._original_pil = self._original_pil.rotate(-degrees, expand=True)
        self._original_pil.load()
        buf = io.BytesIO()
        self._original_pil.save(buf, format="PNG")
        buf.seek(0)
        img = QImage()
        img.loadFromData(buf.getvalue())
        if img.isNull():
            return
        px = QPixmap.fromImage(img)
        # Clear measurements (they'd be in wrong positions after rotation)
        self.clear_all()
        self._scene.removeItem(self.pixmap_item)
        self.pixmap_item = self._scene.addPixmap(px)
        self.pixmap_item.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
        self._scene.setSceneRect(QRectF(px.rect()))
        self.fitInView(self.pixmap_item, Qt.AspectRatioMode.KeepAspectRatio)

    # ── Image loading ─────────────────────────────────────────────────────

    def load_image(self, path: str) -> bool:
        reader = QImageReader(path)
        reader.setAutoTransform(True)
        img = reader.read()
        if img.isNull():
            # Fallback: try loading with Pillow (handles PLU, PLUX and other formats)
            try:
                from PIL import Image
                pil_img = Image.open(path)
                pil_img.load()
                buf = io.BytesIO()
                pil_img.convert("RGBA").save(buf, format="PNG")
                buf.seek(0)
                img = QImage()
                img.loadFromData(buf.getvalue())
                if img.isNull():
                    return False
            except Exception:
                return False
        px = QPixmap.fromImage(img)
        self._reset_state()
        self.pixmap_item = self._scene.addPixmap(px)
        self.pixmap_item.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
        self.setSceneRect(QRectF(px.rect()))
        self.fitInView(self.pixmap_item, Qt.AspectRatioMode.KeepAspectRatio)
        try:
            from PIL import Image
            self._original_pil = Image.open(path)
            self._original_pil.load()
        except Exception:
            self._original_pil = None
        return True

    def _reset_state(self):
        self._scene.clear()
        self.measurements.clear()
        self.measurement_items.clear()
        self._active_count = None
        self._active_count_item = None
        self._scale_bar_item = None   # cleared by _scene.clear()
        self._next_id = 1
        self._color_idx = 0
        self._in_progress.clear()
        self._previews.clear()
        self.pixmap_item = None

    # ── Tool management ───────────────────────────────────────────────────

    def set_tool(self, tool: str):
        if self.current_tool in (Tool.COUNT, Tool.PITS):
            self._finalize_count()
        self._cancel_in_progress()
        self.current_tool = tool
        if tool == Tool.SELECT:
            self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
            self.unsetCursor()   # let ScrollHandDrag manage the hand cursor
        else:
            self.setDragMode(QGraphicsView.DragMode.NoDrag)
            self.setCursor(Qt.CursorShape.CrossCursor)

    def _cancel_in_progress(self):
        self._in_progress.clear()
        for item in self._previews:
            self._scene.removeItem(item)
        self._previews.clear()

    # ── Mouse events ──────────────────────────────────────────────────────

    def mousePressEvent(self, event):
        if self.current_tool == Tool.SELECT:
            super().mousePressEvent(event)
            if event.button() == Qt.MouseButton.LeftButton:
                for it in self.items(event.pos()):
                    if isinstance(it, MeasurementItem):
                        self.measurement_clicked.emit(it.measurement.id)
                        break
            return
        if event.button() == Qt.MouseButton.LeftButton:
            self._handle_click(self._to_scene(event))
        elif event.button() == Qt.MouseButton.RightButton:
            if self.current_tool == Tool.AREA and len(self._in_progress) >= 3:
                self._complete_area()
            elif self.current_tool in (Tool.COUNT, Tool.PITS) and self._active_count:
                self._finalize_count()
            else:
                self._cancel_in_progress()

    def mouseDoubleClickEvent(self, event):
        if self.current_tool == Tool.AREA and len(self._in_progress) >= 3:
            self._complete_area()
        else:
            super().mouseDoubleClickEvent(event)

    def mouseMoveEvent(self, event):
        sp = self.mapToScene(event.pos())
        if self.pixmap_item and self.pixmap_item.boundingRect().contains(sp):
            self.position_changed.emit(sp.x(), sp.y())
        if self.current_tool not in (Tool.SELECT, Tool.COUNT, Tool.PITS) and self._in_progress:
            self._update_preview((sp.x(), sp.y()))
        super().mouseMoveEvent(event)

    def wheelEvent(self, event):
        factor = 1.15 if event.angleDelta().y() > 0 else 1 / 1.15
        self.scale(factor, factor)

    def _to_scene(self, event) -> Tuple[float, float]:
        sp = self.mapToScene(event.pos())
        return (sp.x(), sp.y())

    # ── Click handling ────────────────────────────────────────────────────

    def _handle_click(self, pos: Tuple[float, float]):
        tool = self.current_tool
        if tool in (Tool.DISTANCE, Tool.CALIBRATE):
            self._in_progress.append(pos)
            if len(self._in_progress) == 2:
                if tool == Tool.CALIBRATE:
                    p1, p2 = self._in_progress
                    px_len = math.sqrt((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)
                    self._cancel_in_progress()
                    self.calibration_line_drawn.emit(px_len)
                else:
                    self._complete_simple(MeasurementType.DISTANCE)
        elif tool == Tool.AREA:
            self._in_progress.append(pos)
        elif tool == Tool.ANGLE:
            self._in_progress.append(pos)
            if len(self._in_progress) == 3:
                self._complete_simple(MeasurementType.ANGLE)
        elif tool == Tool.COUNT:
            self._handle_count_click(pos, MeasurementType.COUNT)
        elif tool == Tool.PITS:
            self._handle_count_click(pos, MeasurementType.PITS)

    def _complete_simple(self, mtype: MeasurementType):
        pts = self._in_progress.copy()
        self._cancel_in_progress()
        self._register(Measurement(
            id=self._next_id,
            mtype=mtype,
            points=pts,
            label=f"{mtype.value} {self._next_id}",
            color=self._next_color(),
        ))

    def _complete_area(self):
        pts = self._in_progress.copy()
        self._cancel_in_progress()
        self._register(Measurement(
            id=self._next_id,
            mtype=MeasurementType.AREA,
            points=pts,
            label=f"Area {self._next_id}",
            color=self._next_color(),
        ))

    def _handle_count_click(self, pos: Tuple[float, float], mtype: MeasurementType = MeasurementType.COUNT):
        if self._active_count is None:
            m = Measurement(
                id=self._next_id,
                mtype=mtype,
                points=[pos],
                label=f"{mtype.value} {self._next_id}",
                color=self._next_color(),
            )
            self._next_id += 1
            self._active_count = m
            item = MeasurementItem(m, self.calibration, self.jaw, self.side,
                                   self.show_labels, self.use_puech_colors)
            self._scene.addItem(item)
            self._active_count_item = item
        else:
            self._active_count.points.append(pos)
            self._active_count_item.prepareGeometryChange()
            self._active_count_item.update()

    def _finalize_count(self):
        if self._active_count and self._active_count.points:
            self.measurements.append(self._active_count)
            self.measurement_items.append(self._active_count_item)
            self.measurement_added.emit(self._active_count)
        elif self._active_count_item:
            self._scene.removeItem(self._active_count_item)
        self._active_count = None
        self._active_count_item = None

    def _register(self, m: Measurement):
        self._next_id += 1
        item = MeasurementItem(m, self.calibration, self.jaw, self.side,
                               self.show_labels, self.use_puech_colors)
        self._scene.addItem(item)
        self.measurements.append(m)
        self.measurement_items.append(item)
        self.measurement_added.emit(m)

    def _next_color(self) -> str:
        c = self.COLORS[self._color_idx % len(self.COLORS)]
        self._color_idx += 1
        return c

    # ── Preview drawing ───────────────────────────────────────────────────

    def _update_preview(self, cur: Tuple[float, float]):
        for item in self._previews:
            self._scene.removeItem(item)
        self._previews.clear()

        dash_pen = QPen(QColor("#22C55E"), 1.5, Qt.PenStyle.DashLine)
        dash_pen.setCosmetic(True)
        dot_pen = QPen(QColor("#22C55E"), 1.5)
        dot_pen.setCosmetic(True)

        pts = self._in_progress
        all_pts = pts + [cur]

        for i in range(len(all_pts) - 1):
            line = self._scene.addLine(
                all_pts[i][0], all_pts[i][1],
                all_pts[i+1][0], all_pts[i+1][1],
                dash_pen,
            )
            self._previews.append(line)

        if self.current_tool == Tool.AREA and len(pts) >= 2:
            close_pen = QPen(QColor("#22C55E"), 1, Qt.PenStyle.DotLine)
            close_pen.setCosmetic(True)
            line = self._scene.addLine(cur[0], cur[1], pts[0][0], pts[0][1], close_pen)
            self._previews.append(line)

        for p in pts:
            r = 4
            circ = self._scene.addEllipse(
                p[0]-r, p[1]-r, 2*r, 2*r, dot_pen, QBrush(Qt.GlobalColor.white)
            )
            self._previews.append(circ)

    # ── Public API ────────────────────────────────────────────────────────

    def delete_by_id(self, mid: int):
        for i, m in enumerate(self.measurements):
            if m.id == mid:
                self._scene.removeItem(self.measurement_items[i])
                self.measurements.pop(i)
                self.measurement_items.pop(i)
                return

    def save_annotated_image(self, path: str) -> bool:
        """Render the image + all measurement overlays to a file (PNG/JPG/TIFF)."""
        if not self.pixmap_item:
            return False
        src = self.pixmap_item.boundingRect()
        img = QImage(src.size().toSize(), QImage.Format.Format_ARGB32)
        img.fill(Qt.GlobalColor.white)
        painter = QPainter(img)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self._scene.render(painter, source=src)
        painter.end()
        return img.save(path)

    def clear_all(self):
        for item in self.measurement_items:
            self._scene.removeItem(item)
        self.measurements.clear()
        self.measurement_items.clear()
        if self._active_count_item:
            self._scene.removeItem(self._active_count_item)
        self._active_count = None
        self._active_count_item = None

    def update_calibration(self, cal: Calibration):
        self.calibration = cal
        for item in self.measurement_items:
            item.calibration = cal
            item.update()
        if self._active_count_item:
            self._active_count_item.calibration = cal
            self._active_count_item.update()

    def fit_view(self):
        if self.pixmap_item:
            self.fitInView(self.pixmap_item, Qt.AspectRatioMode.KeepAspectRatio)

    def zoom_in(self):
        self.scale(1.25, 1.25)

    def zoom_out(self):
        self.scale(0.8, 0.8)

    def zoom_reset(self):
        self.resetTransform()
        self.fit_view()

    def image_size(self) -> Tuple[int, int]:
        if self.pixmap_item:
            px = self.pixmap_item.pixmap()
            return px.width(), px.height()
        return 0, 0

    # ── Multi-image state ─────────────────────────────────────────────────

    def get_state(self) -> dict:
        return {
            "measurements": list(self.measurements),
            "calibration":  self.calibration,
            "next_id":      self._next_id,
            "color_idx":    self._color_idx,
            "jaw":          self.jaw,
            "side":         self.side,
            "tooth":        self.tooth,
        }

    def restore_state(self, state: dict):
        for item in self.measurement_items:
            self._scene.removeItem(item)
        self.measurement_items.clear()
        self.measurements  = list(state["measurements"])
        self.calibration   = state["calibration"]
        self._next_id      = state["next_id"]
        self._color_idx    = state["color_idx"]
        self.jaw           = state.get("jaw",   "Upper")
        self.side          = state.get("side",  "Right")
        self.tooth         = state.get("tooth", "I1")
        for m in self.measurements:
            item = MeasurementItem(m, self.calibration, self.jaw, self.side,
                                   self.show_labels, self.use_puech_colors)
            self._scene.addItem(item)
            self.measurement_items.append(item)

    def set_puech_colors(self, enabled: bool):
        self.use_puech_colors = enabled
        for item in self.measurement_items:
            item.use_puech_colors = enabled
            item.update()
        if self._active_count_item:
            self._active_count_item.use_puech_colors = enabled
            self._active_count_item.update()

    def set_labels_visible(self, visible: bool):
        self.show_labels = visible
        for item in self.measurement_items:
            item.show_labels = visible
            item.update()
        if self._active_count_item:
            self._active_count_item.show_labels = visible
            self._active_count_item.update()

    def set_tooth_context(self, jaw: str, side: str, tooth: str):
        self.jaw   = jaw
        self.side  = side
        self.tooth = tooth
        for item in self.measurement_items:
            item.jaw  = jaw
            item.side = side
            item.update()
        if self._active_count_item:
            self._active_count_item.jaw  = jaw
            self._active_count_item.side = side
            self._active_count_item.update()

    # ── Filters ───────────────────────────────────────────────────────────

    def apply_filters(self, brightness=0, contrast=0, sharpness=0, blur=0,
                      grayscale=False, invert=False, special=None):
        if self._original_pil is None or self.pixmap_item is None:
            return
        try:
            from PIL import Image, ImageEnhance, ImageFilter, ImageOps
        except ImportError:
            return
        img = self._original_pil.copy().convert("RGB")

        if special == "auto":
            img = ImageOps.autocontrast(img)
        elif special == "equalize":
            img = ImageOps.equalize(img)
        elif special == "edge":
            img = img.filter(ImageFilter.FIND_EDGES)
        elif special == "sharpen":
            img = img.filter(ImageFilter.SHARPEN)

        if grayscale:
            img = ImageOps.grayscale(img).convert("RGB")
        if invert:
            img = ImageOps.invert(img)
        if brightness != 0:
            img = ImageEnhance.Brightness(img).enhance(1.0 + brightness / 100.0)
        if contrast != 0:
            img = ImageEnhance.Contrast(img).enhance(1.0 + contrast / 100.0)
        if sharpness != 0:
            img = ImageEnhance.Sharpness(img).enhance(1.0 + sharpness / 100.0)
        if blur > 0:
            img = img.filter(ImageFilter.GaussianBlur(radius=blur / 3.0))

        self._pil_to_canvas(img)

    def reset_filters(self):
        if self._original_pil is None or self.pixmap_item is None:
            return
        img = self._original_pil.copy().convert("RGB")
        self._pil_to_canvas(img)

    def _pil_to_canvas(self, img):
        data = img.tobytes("raw", "RGB")
        qimg = QImage(data, img.width, img.height, img.width * 3,
                      QImage.Format.Format_RGB888)
        self.pixmap_item.setPixmap(QPixmap.fromImage(qimg))

    # ── Export ────────────────────────────────────────────────────────────

    def _row_data(self, m):
        return self._row_data_with_cal(m, self.calibration)

    def _row_data_with_cal(self, m, cal, jaw="Upper", side="Right", tooth=""):
        pv = m.pixel_value()
        if m.mtype == MeasurementType.COUNT:
            val, unit = str(len(m.points)), "objects"
        elif m.mtype == MeasurementType.PITS:
            val, unit = str(len(m.points)), "pits"
        elif m.mtype == MeasurementType.ANGLE:
            val, unit = f"{pv:.4f}", "degrees"
        elif m.mtype == MeasurementType.AREA:
            val = f"{cal.to_real_area(pv):.4f}"
            unit = cal.unit_str(is_area=True)
        else:
            val = f"{cal.to_real(pv):.4f}"
            unit = cal.unit_str()
        a = m.slope_angle()
        slope  = f"{abs(a):.2f}" if a is not None else "-"
        puech  = str(m.puech(jaw, side)) if a is not None else "-"
        pts    = "; ".join(f"({p[0]:.1f},{p[1]:.1f})" for p in m.points)
        stype  = str(m.striation_type) if m.striation_type else "-"
        return [m.id, m.mtype.value, tooth, jaw, side, unit, val, slope, puech, stype, pts]

    def export_csv(self, path: str):
        import csv
        headers = ["ID", "Type", "Value", "Unit", "Slope (\u00b0)", "Label", "Points (px)"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for m in self.measurements:
                w.writerow(self._row_data(m))

    def export_txt(self, path: str):
        headers = ["ID", "Type", "Value", "Unit", "Slope (\u00b0)", "Label", "Points (px)"]
        with open(path, "w", encoding="utf-8") as f:
            f.write("\t".join(headers) + "\n")
            for m in self.measurements:
                f.write("\t".join(str(x) for x in self._row_data(m)) + "\n")

    def export_xlsx(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OpenScan Measurements"

        headers = ["ID", "Type", "Value", "Unit", "Slope (\u00b0)", "Label", "Points (px)"]
        hdr_fill = PatternFill("solid", fgColor="4ADE80")
        hdr_font = Font(bold=True, color="14532D")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for row, m in enumerate(self.measurements, 2):
            for col, val in enumerate(self._row_data(m), 1):
                ws.cell(row=row, column=col, value=val)

        for col_cells in ws.columns:
            width = max(len(str(c.value or "")) for c in col_cells) + 4
            ws.column_dimensions[col_cells[0].column_letter].width = min(width, 45)

        # Zebra striping
        alt_fill = PatternFill("solid", fgColor="ECFDF5")
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill

        wb.save(path)
